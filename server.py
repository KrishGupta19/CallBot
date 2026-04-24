import os
import sys
import json
import asyncio
import io
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from loguru import logger
import aiohttp
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, WebSocket, Request
from fastapi.responses import RedirectResponse
from fastapi.responses import PlainTextResponse, HTMLResponse
from fastapi.responses import Response, StreamingResponse
from twilio.rest import Client as TwilioClient

load_dotenv(override=True)

# Configure logging
try:
    logger.remove(0)
except ValueError:
    pass
logger.add(sys.stderr, level="DEBUG")

app = FastAPI()

NGROK_URL = os.getenv("NGROK_URL", "").replace("https://", "").replace("http://", "").rstrip("/")

LOGS_DIR = Path(__file__).parent / "call_logs"
RECORDINGS_DIR = Path(__file__).parent / "call_recordings"
RECORDINGS_INDEX_PATH = LOGS_DIR / "_recordings_index.json"


def _load_recordings_index() -> dict:
    try:
        if not RECORDINGS_INDEX_PATH.exists():
            return {}
        data = json.loads(RECORDINGS_INDEX_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_recordings_index(index: dict) -> None:
    try:
        LOGS_DIR.mkdir(exist_ok=True)
        RECORDINGS_INDEX_PATH.write_text(
            json.dumps(index, indent=2, ensure_ascii=False), encoding="utf-8"
        )
    except Exception as e:
        logger.warning(f"Failed to write recordings index: {e}")


def _attach_recording_from_index(call_data: dict, index: dict) -> dict:
    if not isinstance(call_data, dict):
        return call_data
    if call_data.get("recording_url"):
        return call_data
    call_sid = call_data.get("call_sid")
    if not call_sid:
        return call_data
    rec = index.get(call_sid)
    if not isinstance(rec, dict):
        return call_data
    recording_sid = rec.get("recording_sid")
    if not recording_sid:
        return call_data
    call_data["recording_sid"] = recording_sid
    call_data["recording_status"] = rec.get("recording_status") or "completed"
    # Prefer local file if present, else proxy endpoint.
    local_path = RECORDINGS_DIR / f"{recording_sid}.mp3"
    call_data["recording_url"] = (
        f"/recordings-local/{recording_sid}.mp3"
        if local_path.exists()
        else f"/recordings/{recording_sid}.mp3"
    )
    return call_data


def _public_base_url(request: Request) -> str:
    """
    Best-effort public base URL (works behind ngrok).
    Twilio requires absolute URLs for action/callback.
    """
    proto = request.headers.get("x-forwarded-proto") or request.url.scheme
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc
    return f"{proto}://{host}".rstrip("/")


def _public_ws_url(request: Request) -> str:
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc
    # Twilio <Stream> must use wss:// for public endpoints.
    return f"wss://{host}/ws"


async def _start_twilio_recording(call_sid: str, *, callback_url: str) -> None:
    """
    Start recording via Twilio REST API (more reliable with <Connect><Stream>).
    """
    account_sid = os.getenv("TWILIO_ACCOUNT_SID") or ""
    auth_token = os.getenv("TWILIO_AUTH_TOKEN") or ""
    if not account_sid or not auth_token:
        logger.warning("TWILIO_ACCOUNT_SID/TWILIO_AUTH_TOKEN not set; cannot start recording")
        return

    def _run():
        client = TwilioClient(account_sid, auth_token)
        rec = client.calls(call_sid).recordings.create(
            recording_status_callback=callback_url,
            recording_status_callback_event=["completed"],
        )
        return rec.sid

    try:
        rec_sid = await asyncio.to_thread(_run)
        logger.info(f"Started Twilio recording: call_sid={call_sid} recording_sid={rec_sid}")
    except Exception as e:
        logger.warning(f"Failed to start Twilio recording via REST: {e}")


@app.get("/health")
async def health_check():
    return {"status": "ok", "service": "ai-voice-agent"}

@app.get("/")
async def root():
    return RedirectResponse(url="/dashboard")

@app.get("/calls")
async def get_calls():
    """Return recent call logs."""
    logs_dir = LOGS_DIR
    if not logs_dir.exists():
        return {"calls": []}
    
    recordings_index = _load_recordings_index()
    calls = []
    for filepath in sorted(
        logs_dir.glob("call_*.json"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )[:20]:
        try:
            call_data = json.loads(filepath.read_text(encoding="utf-8"))
            if isinstance(call_data, list):
                # Legacy format: transcript-only list (sometimes empty). Inject call_start
                # from file modified time so the dashboard can show date/time.
                mtime = datetime.fromtimestamp(filepath.stat().st_mtime).isoformat()
                calls.append(
                    {
                        "filename": filepath.name,
                        "call_start": mtime,
                        "call_end": None,
                        "duration_seconds": None,
                        "caller_number": None,
                        "business_type": "unknown",
                        "transcript": call_data,
                        "transcript_text": " ".join(
                            [
                                (t.get("text") or "").strip()
                                for t in call_data
                                if isinstance(t, dict)
                            ]
                        ).strip(),
                        "intake": None,
                        "summary": {
                            "caller_name": None,
                            "interest": None,
                            "budget_range": None,
                            "timeline": None,
                            "source": None,
                            "lead_score": None,
                            "sentiment": None,
                            "summary": None,
                            "follow_up_action": None,
                        },
                    }
                )
            else:
                call_data["filename"] = filepath.name
                call_data.setdefault("business_type", "unknown")
                call_data.setdefault("intake", None)
                call_data = _attach_recording_from_index(call_data, recordings_index)
                calls.append(call_data)
        except Exception as e:
            logger.error(f"Failed to read {filepath}: {e}")
    
    return {"calls": calls, "total": len(calls)}


@app.get("/calls/{filename}")
async def get_call_detail(filename: str):
    """Return a specific call log."""
    filepath = LOGS_DIR / filename
    if not filepath.exists():
        return {"error": "Call not found"}
    data = json.loads(filepath.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = _attach_recording_from_index(data, _load_recordings_index())
    return data


def _flatten_call_for_export(call_data: dict, *, recordings_index: dict, file_mtime_iso: str | None) -> dict:
    call_data = _attach_recording_from_index(dict(call_data), recordings_index)

    summary = call_data.get("summary") if isinstance(call_data.get("summary"), dict) else {}
    intake = call_data.get("intake") if isinstance(call_data.get("intake"), dict) else {}

    bt = (call_data.get("business_type") or (intake.get("business_type") if isinstance(intake, dict) else "") or "unknown").strip()
    outcome = (intake.get("outcome") or "").strip()

    lead_score = summary.get("lead_score")
    sentiment = summary.get("sentiment")

    items = intake.get("items")
    patient_name = intake.get("patient_name")
    customer_name = intake.get("customer_name")

    return {
        "filename": call_data.get("filename") or "",
        "call_sid": call_data.get("call_sid") or "",
        "business_type": bt,
        "caller_number": call_data.get("caller_number") or "",
        "call_start": call_data.get("call_start") or file_mtime_iso or "",
        "call_end": call_data.get("call_end") or "",
        "duration_seconds": call_data.get("duration_seconds") if call_data.get("duration_seconds") is not None else "",
        "lead_score": lead_score if lead_score is not None else "",
        "sentiment": sentiment if sentiment is not None else "",
        "outcome": outcome,
        "patient_name": patient_name or "",
        "customer_name": customer_name or "",
        "items": items or "",
        "recording_url": call_data.get("recording_url") or "",
        "transcript_text": call_data.get("transcript_text") or "",
        "intake_json": json.dumps(intake, ensure_ascii=False) if intake else "",
        "summary_text": summary.get("summary") or "",
        "follow_up_action": summary.get("follow_up_action") or "",
    }


@app.get("/export/excel")
async def export_excel():
    """Export ALL call logs to an .xlsx file."""
    LOGS_DIR.mkdir(exist_ok=True)
    recordings_index = _load_recordings_index()

    rows: list[dict] = []
    for fp in sorted(LOGS_DIR.glob("call_*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            raw = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue

        mtime_iso = datetime.fromtimestamp(fp.stat().st_mtime).isoformat()

        if isinstance(raw, list):
            # Legacy transcript-only format.
            transcript_text = " ".join(
                [(t.get("text") or "").strip() for t in raw if isinstance(t, dict)]
            ).strip()
            data = {
                "filename": fp.name,
                "call_sid": "",
                "business_type": "unknown",
                "caller_number": "",
                "call_start": mtime_iso,
                "call_end": "",
                "duration_seconds": "",
                "transcript_text": transcript_text,
                "intake": None,
                "summary": {"lead_score": "", "sentiment": "", "summary": "", "follow_up_action": ""},
            }
            rows.append(_flatten_call_for_export(data, recordings_index=recordings_index, file_mtime_iso=mtime_iso))
        elif isinstance(raw, dict):
            raw = dict(raw)
            raw.setdefault("filename", fp.name)
            raw.setdefault("business_type", "unknown")
            raw.setdefault("intake", None)
            raw.setdefault("summary", {})
            rows.append(_flatten_call_for_export(raw, recordings_index=recordings_index, file_mtime_iso=mtime_iso))

    wb = Workbook()
    ws = wb.active
    ws.title = "Calls"

    cols = [
        "filename",
        "call_sid",
        "business_type",
        "caller_number",
        "call_start",
        "call_end",
        "duration_seconds",
        "lead_score",
        "sentiment",
        "outcome",
        "patient_name",
        "customer_name",
        "items",
        "recording_url",
        "transcript_text",
        "intake_json",
        "summary_text",
        "follow_up_action",
    ]

    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    # Basic formatting: bold header and reasonable column widths
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(12, len(c) + 2))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"calls_{stamp}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/dashboard")
async def dashboard():
    """Serve the web dashboard."""
    dashboard_path = Path(__file__).parent / "dashboard.html"
    if not dashboard_path.exists():
        return HTMLResponse("<h1>Dashboard not found</h1><p>Please create dashboard.html</p>", status_code=404)
    return HTMLResponse(dashboard_path.read_text(encoding="utf-8"))

@app.get("/dashboard/")
async def dashboard_slash():
    return RedirectResponse(url="/dashboard")


@app.post("/incoming-call")
async def incoming_call(request: Request):
    """
    Twilio hits this endpoint when someone calls your number.
    We return TwiML with a simple IVR menu:
    - Press 1 for doctor appointments
    - Press 2 for bakery orders
    """
    form_data = await request.form()
    caller_number = form_data.get("From", "unknown")
    logger.info(f"Incoming call from: {caller_number}")

    public_base = _public_base_url(request)
    route_url = f"{public_base}/route-call"

    # Return the IVR menu
    twiml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Gather input="dtmf" numDigits="1" timeout="7" action="{route_url}" method="POST">
        <Say voice="alice">Welcome. For doctor appointments, press 1. For bakery orders, press 2.</Say>
    </Gather>
    <Say voice="alice">Sorry, I did not receive your selection.</Say>
    <Redirect method="POST">{public_base}/incoming-call</Redirect>
</Response>"""

    return PlainTextResponse(content=twiml, media_type="application/xml")


@app.post("/route-call")
async def route_call(request: Request):
    """Handle IVR digit and start streaming to /ws with business_type."""
    form_data = await request.form()
    digits = (form_data.get("Digits") or "").strip()
    caller_number = form_data.get("From", "unknown")
    call_sid = (form_data.get("CallSid") or "").strip()

    business_type = "doctor" if digits == "1" else "bakery" if digits == "2" else "unknown"
    logger.info(f"IVR selection digits={digits!r} -> business_type={business_type} caller={caller_number}")

    public_base = _public_base_url(request)
    ws_url = _public_ws_url(request) if not NGROK_URL else f"wss://{NGROK_URL}/ws"

    if business_type == "unknown":
        twiml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Invalid selection.</Say>
    <Redirect method="POST">{public_base}/incoming-call</Redirect>
</Response>"""
        return PlainTextResponse(content=twiml, media_type="application/xml")

    recording_status_url = f"{public_base}/recording-status"
    # Start recording as early as possible after IVR selection.
    # Starting here avoids delays from websocket/bot initialization.
    if call_sid:
        asyncio.create_task(_start_twilio_recording(call_sid, callback_url=recording_status_url))

    twiml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Start>
        <Record recordingStatusCallback="{recording_status_url}" recordingStatusCallbackMethod="POST" recordingStatusCallbackEvent="completed" />
    </Start>
    <Connect>
        <Stream url="{ws_url}">
            <Parameter name="caller_number" value="{caller_number}" />
            <Parameter name="business_type" value="{business_type}" />
        </Stream>
    </Connect>
</Response>"""
    return PlainTextResponse(content=twiml, media_type="application/xml")


@app.post("/recording-status")
async def recording_status(request: Request):
    """
    Twilio calls this when a recording is available.
    We attach recording info to the matching call log by call_sid.
    """
    form_data = await request.form()
    call_sid = (form_data.get("CallSid") or "").strip()
    recording_sid = (form_data.get("RecordingSid") or "").strip()
    recording_status = (form_data.get("RecordingStatus") or "").strip()

    logger.info(f"Recording status: call_sid={call_sid} recording_sid={recording_sid} status={recording_status}")

    if not call_sid or not recording_sid:
        return PlainTextResponse("missing CallSid/RecordingSid", status_code=400)

    logs_dir = LOGS_DIR
    if not logs_dir.exists():
        logs_dir.mkdir(exist_ok=True)

    # Persist mapping even if the call log isn't written yet (callback timing).
    index = _load_recordings_index()
    index[call_sid] = {
        "recording_sid": recording_sid,
        "recording_status": recording_status or "completed",
        "updated_at": datetime.now().isoformat(),
    }
    _save_recordings_index(index)

    # Find the most recent log matching this call_sid
    candidates = sorted(logs_dir.glob("call_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)[:80]
    for fp in candidates:
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        if isinstance(data, dict) and (data.get("call_sid") == call_sid):
            data["recording_sid"] = recording_sid
            # Prefer serving from local if downloaded, else proxy endpoint.
            local_path = RECORDINGS_DIR / f"{recording_sid}.mp3"
            data["recording_url"] = (
                f"/recordings-local/{recording_sid}.mp3"
                if local_path.exists()
                else f"/recordings/{recording_sid}.mp3"
            )
            data["recording_status"] = recording_status or "completed"
            fp.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
            break

    # Best-effort: download recording locally for long-term storage.
    try:
        RECORDINGS_DIR.mkdir(exist_ok=True)
        await _download_recording_mp3(recording_sid)
    except Exception as e:
        logger.warning(f"Failed to download recording locally: {e}")

    return PlainTextResponse("ok")

async def _download_recording_mp3(recording_sid: str) -> None:
    account_sid = os.getenv("TWILIO_ACCOUNT_SID") or ""
    auth_token = os.getenv("TWILIO_AUTH_TOKEN") or ""
    if not account_sid or not auth_token:
        raise RuntimeError("Missing TWILIO_ACCOUNT_SID/TWILIO_AUTH_TOKEN")

    out_path = RECORDINGS_DIR / f"{recording_sid}.mp3"
    if out_path.exists() and out_path.stat().st_size > 0:
        return

    url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Recordings/{recording_sid}.mp3"
    auth = aiohttp.BasicAuth(login=account_sid, password=auth_token)
    async with aiohttp.ClientSession(auth=auth) as session:
        async with session.get(url) as resp:
            if resp.status >= 400:
                text = await resp.text()
                raise RuntimeError(f"Twilio download failed: {resp.status} {text[:200]}")
            out_path.write_bytes(await resp.read())


@app.get("/recordings/{recording_sid}.mp3")
async def recording_proxy_mp3(recording_sid: str):
    """
    Proxy Twilio recording audio to the browser (Twilio requires auth).
    Requires TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN in env.
    """
    account_sid = os.getenv("TWILIO_ACCOUNT_SID") or ""
    auth_token = os.getenv("TWILIO_AUTH_TOKEN") or ""
    if not account_sid or not auth_token:
        return PlainTextResponse("Twilio credentials not configured", status_code=500)

    # Twilio recording media URL pattern:
    # https://api.twilio.com/2010-04-01/Accounts/{AccountSid}/Recordings/{RecordingSid}.mp3
    url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Recordings/{recording_sid}.mp3"

    async def iter_bytes():
        auth = aiohttp.BasicAuth(login=account_sid, password=auth_token)
        async with aiohttp.ClientSession(auth=auth) as session:
            async with session.get(url) as resp:
                if resp.status >= 400:
                    text = await resp.text()
                    raise RuntimeError(f"Twilio fetch failed: {resp.status} {text[:200]}")
                async for chunk in resp.content.iter_chunked(1024 * 64):
                    yield chunk

    try:
        return StreamingResponse(iter_bytes(), media_type="audio/mpeg")
    except Exception as e:
        logger.error(f"Recording proxy error: {e}")
        return PlainTextResponse("Failed to fetch recording", status_code=502)


@app.get("/recordings-local/{recording_sid}.mp3")
async def recording_local_mp3(recording_sid: str):
    """Serve a previously downloaded recording from disk."""
    path = RECORDINGS_DIR / f"{recording_sid}.mp3"
    if not path.exists():
        return PlainTextResponse("Recording not found locally", status_code=404)
    return Response(content=path.read_bytes(), media_type="audio/mpeg")

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    logger.info("WebSocket connection accepted from Twilio")

    # Read messages from Twilio until we get the 'start' event with stream_sid
    stream_sid = None
    call_sid = None
    caller_number = None
    business_type = None
    try:
        async for message in websocket.iter_text():
            data = json.loads(message)
            event = data.get("event")
            if event == "connected":
                logger.info("Twilio stream connected")
                continue
            elif event == "start":
                stream_sid = data["start"]["streamSid"]
                call_sid = data["start"].get("callSid")
                custom = data["start"].get("customParameters") or {}
                caller_number = custom.get("caller_number") or custom.get("From") or None
                business_type = custom.get("business_type") or None
                logger.info(f"Stream started: stream_sid={stream_sid}, call_sid={call_sid}")
                break
    except Exception as e:
        logger.error(f"Error reading Twilio start event: {e}")
        await websocket.close()
        return

    if not stream_sid:
        logger.error("Never received Twilio stream_sid, closing connection")
        await websocket.close()
        return

    try:
        from bot import run_bot
        await run_bot(websocket, stream_sid, call_sid, caller_number, business_type)
    except asyncio.CancelledError:
        logger.info("Call cancelled (caller hung up)")
    except ConnectionError as e:
        logger.warning(f"Connection lost: {e}")
    except Exception as e:
        logger.error(f"Unexpected bot error: {e}", exc_info=True)
    finally:
        try:
            await websocket.close()
        except Exception:
            pass
        logger.info("WebSocket cleanup complete")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "server:app",
        host="0.0.0.0",
        port=8765,
        log_level="info",
    )
